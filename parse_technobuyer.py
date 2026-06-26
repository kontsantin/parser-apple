"""
Парсер товаров Apple → Яндекс Кит (XLSX) + Яндекс Маркет (YML)
Использование:
  python parse_technobuyer.py "URL-товара"         — автосбор вариантов с одного URL
  python parse_technobuyer.py --file urls.txt      — парсинг списка URL из файла
"""
import sys, re, json, os, html as html_mod, threading
from datetime import datetime
from urllib.request import urlopen, Request
from urllib.parse import urljoin
from concurrent.futures import ThreadPoolExecutor, as_completed
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom.minidom import parseString

import openpyxl

BASE = "https://techno-buyer.ru"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}


import socket
import time as time_mod

socket.setdefaulttimeout(30)

RETRIES = 3
RETRY_DELAY = 3

def fetch(url, attempt=1):
    try:
        req = Request(url, headers=HEADERS)
        with urlopen(req, timeout=30) as resp:
            return resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        if attempt < RETRIES:
            print(f"  [retry {attempt}/{RETRIES}] {url}: {e}")
            time_mod.sleep(RETRY_DELAY * attempt)
            return fetch(url, attempt + 1)
        raise


def parse_product_page(url):
    text = fetch(url)

    price_match = re.search(r'<meta\s+itemprop="price"\s+content="([^"]+)"', text)
    price = float(price_match.group(1)) if price_match else 0

    cp_match = re.search(r'"compare_price":"([^"]+)"', text)
    old_price = float(cp_match.group(1)) if cp_match else 0

    sku_match = re.search(r'<meta\s+itemprop="sku"\s+content="([^"]+)"', text)
    sku = sku_match.group(1) if sku_match else ""

    pid_match = re.search(r'"product_id":"(\d+)"', text)
    product_id = pid_match.group(1) if pid_match else ""

    stock_match = re.search(r'data-sku-count="(\d+)"', text)
    stock = int(stock_match.group(1)) if stock_match else 0

    name_match = re.search(r'<h1[^>]*>([^<]+)</h1>', text)
    name = html_mod.unescape(name_match.group(1).strip()) if name_match else ""

    features = {}
    for m in re.finditer(
        r'<div class="features-two__name"><span>([^<]+)</span></div>\s*<div class="features-two__value">(.*?)</div>',
        text, re.DOTALL
    ):
        key = html_mod.unescape(m.group(1).strip())
        val = re.sub(r'<[^>]+>', '', m.group(2)).strip()
        val = html_mod.unescape(val)
        features[key] = val

    images = []
    seen_img_urls = set()
    for m in re.finditer(r'href="(/wa-data/public/shop/products/.*?\.(?:jpg|png))"', text):
        img_url = urljoin(BASE, m.group(1))
        if img_url in seen_img_urls:
            continue
        seen_img_urls.add(img_url)
        images.append(img_url)

    desc_match = re.search(r'<meta\s+name="description"\s+content="([^"]+)"', text)
    desc = html_mod.unescape(desc_match.group(1)) if desc_match else ""

    brand_match = re.search(r'<meta\s+itemprop="name"\s+content="([^"]+)"', text)
    brand = brand_match.group(1) if brand_match else "Apple"

    SIZE_RE = re.compile(r'^\d+', re.IGNORECASE)

    def _model_name(slug):
        parts = slug.split("-")
        for i, p in enumerate(parts):
            if SIZE_RE.match(p) and any(p.lower().endswith(s) for s in ('gb', 'tb', 'mm')):
                return "-".join(parts[:i])
        return slug

    base_slug_match = re.search(r'/([^/]+)/?(?:index\.php)?$', url)
    model_name = _model_name(base_slug_match.group(1)) if base_slug_match else ""

    variant_urls = set()
    for m in re.finditer(r'<a\s+[^>]*class="[^"]*\bproduct-group__item\b[^"]*"[^>]*href="([^"]+)"', text):
        href = m.group(1)
        if not href:
            continue
        if model_name:
            v_slug_match = re.search(r'/([^/]+)/?(?:index\.php)?$', href)
            if not v_slug_match or _model_name(v_slug_match.group(1)) != model_name:
                continue
        variant_urls.add(urljoin(BASE, href))

    return {
        "url": url,
        "name": name,
        "price": price,
        "old_price": old_price,
        "sku": sku,
        "product_id": product_id,
        "stock": stock,
        "brand": brand,
        "features": features,
        "images": images,
        "description": desc,
        "variant_urls": variant_urls,
        "source_url": url,
    }


def parse_product(main_url):
    print(f"Парсинг главной страницы...")
    main = parse_product_page(main_url)
    main["source_url"] = main_url
    variants = [main]
    parsed_urls = {main_url}

    for vurl in sorted(main["variant_urls"]):
        if vurl in parsed_urls:
            continue
        parsed_urls.add(vurl)
        try:
            print(f"  Парсинг варианта: {vurl}")
            v = parse_product_page(vurl)
            v["source_url"] = main_url
            variants.append(v)
        except Exception as e:
            print(f"  [!] Ошибка: {e}")

    return variants


def detect_category(url):
    u = url.lower()
    if "airpods" in u:
        return "airpods"
    if "apple-watch" in u or "/watch" in u:
        return "apple_watch"
    if "macbook" in u:
        return "macbook"
    if "ipad" in u:
        return "ipad"
    if "iphone" in u:
        return "iphone"
    return "other"


# --- Error/debug tracking ---

def _errors_path(urls_path):
    base = os.path.splitext(urls_path)[0]
    return f"{base}_errors.txt"


def load_error_urls(errors_path):
    if not os.path.exists(errors_path):
        return []
    with open(errors_path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip() and not line.strip().startswith("#")]


def save_error_url(errors_path, url):
    existing = set()
    if os.path.exists(errors_path):
        with open(errors_path, "r", encoding="utf-8") as f:
            existing = {line.strip() for line in f if line.strip()}
    if url not in existing:
        with open(errors_path, "a", encoding="utf-8") as f:
            f.write(url + "\n")


def parse_urls_from_file(filepath, max_workers=10):
    errors_path = _errors_path(filepath)

    # 1. Load main URLs
    with open(filepath, "r", encoding="utf-8") as f:
        urls = [line.strip() for line in f if line.strip() and not line.strip().startswith("#")]

    # 2. Load previously failed URLs
    retry_urls = load_error_urls(errors_path)
    if retry_urls:
        print(f"Найдено URL с прошлых ошибок: {len(retry_urls)}")
        existing = set(urls)
        retry_only = [u for u in retry_urls if u not in existing]
        urls = retry_only + urls
        # Clear old error file — will repopulate with only still-failing URLs
        open(errors_path, "w", encoding="utf-8").close()

    if not urls:
        raise SystemExit(f"Файл {filepath} пуст или не содержит URL.")

    print(f"Загружено URL: {len(urls)} (потоков: {max_workers})")
    variants = []
    parsed_urls = set(urls)
    lock = threading.Lock()
    total = len(urls)
    done = [0]
    failed_urls_lock = threading.Lock()
    failed_urls = set()
    failed_variant_urls = set()

    def _process_url(url):
        local_variants = []
        try:
            main = parse_product_page(url)
            main["source_url"] = url
            with lock:
                done[0] += 1
                idx = done[0]
            print(f"\n[{idx}/{total}] Парсинг: {url}")
            if main["url"] not in parsed_urls:
                with lock:
                    if main["url"] not in parsed_urls:
                        parsed_urls.add(main["url"])
                        local_variants.append(main)

            for vurl in sorted(main["variant_urls"]):
                with lock:
                    if vurl in parsed_urls:
                        continue
                    parsed_urls.add(vurl)
                try:
                    print(f"  Вариант: {vurl}")
                    v = parse_product_page(vurl)
                    v["source_url"] = url
                    local_variants.append(v)
                except Exception as e:
                    print(f"  [!] Ошибка варианта: {e}")
                    with failed_urls_lock:
                        failed_variant_urls.add(vurl)
        except Exception as e:
            with lock:
                done[0] += 1
            print(f"  [!] Ошибка [{url}]: {e}")
            with failed_urls_lock:
                failed_urls.add(url)
        return local_variants

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_process_url, url): url for url in urls}
        for future in as_completed(futures):
            result = future.result()
            with lock:
                variants.extend(result)

    # 3. Write failed URLs to error file for next retry
    all_failed = failed_urls | failed_variant_urls
    if all_failed:
        with open(errors_path, "w", encoding="utf-8") as f:
            for u in sorted(all_failed):
                f.write(u + "\n")
        print(f"\n  [!] Не спарсилось URL: {len(all_failed)} — записаны в {errors_path}")
    elif retry_urls:
        # All previous errors resolved — remove file
        if os.path.exists(errors_path):
            os.remove(errors_path)

    return variants


# ====================================================================
# YANDEX KIT XLSX GENERATION
# ====================================================================

def _storage_label(val):
    m = re.search(r'(\d+)', val)
    if not m:
        return val
    gb = int(m.group(1))
    if gb == 1024:
        return "1Tb"
    return f"{gb}Gb"


def _color_label(val):
    return val.strip().capitalize()


def _build_product_name(variant, features, series):
    color = features.get("Цвет", "")
    memory = features.get("Встроенная память", "")
    conn = features.get("Связь", "")
    base = series
    if memory:
        base += " " + _storage_label(memory)
    if conn:
        base += ", " + conn
    base += ", без RuStore"
    if color:
        base += " (" + _color_label(color) + ")"
    return base


def generate_yandex_kit_xlsx(variants, output_path):

    if not variants:
        return 0

    # Feature keys used by the 12 fixed characteristic columns
    FIXED_FEATURE_KEYS = {
        "Цвет", "Встроенная память", "Связь", "Серия", "Процессор", "Диагональ",
        "Разрешение камеры", "Разрешение фронтальной камеры, Мп", "Операционная система",
        "Защита от воды", "Разъём", "Год (модель представлена)", "Бренд",
    }

    # Collect extra characteristic columns: all features not in fixed columns
    EXTRA_PRIORITY = [
        "Тип",
        "Ядер процессора", "Ядер Neural Engine", "Ядер графического процессора",
        "Оперативная память", "Тип накопителя", "Поддержка карт памяти",
        "Тип дисплея", "Разрешение экрана, пикс",
        "Технологии дисплея", "Плотность пикселей на дюйм", "Яркость, кд/м²",
        "Контрастность", "Цветовой охват", "Поддержка доп. мониторов",
        "Беспроводная сеть", "Сотовая и беспроводная сеть",
        "Поддержка интерфейсов", "Количество HDMI", "Количество Thunderbolt/USB 4",
        "Диафрагма", "Зум (фото)",
        "Разрешение видео", "Разрешение замедленного видео",
        "Разрешение видео фронтальной камеры",
        "Функции камеры", "Функции фронтальной камеры", "Функции видео",
        "Защита объектива", "Веб-камера",
        "Количество микрофонов", "Трекпад", "Аудио",
        "Датчики", "Навигация",
        "Тип аккумулятора", "Работа от аккумулятора, часов",
        "Время работы", "Мощность адаптера", "Разъем питания",
        "Материал", "Вес", "Размер", "Ширина, мм", "Высота, мм", "Длина, мм",
        "Гарантия", "Страна производства", "В комплекте",
    ]

    extra_keys = []
    seen_extra = set()
    for key in EXTRA_PRIORITY:
        if key not in FIXED_FEATURE_KEYS and key not in seen_extra:
            seen_extra.add(key)
            extra_keys.append(key)
    for v in variants:
        for key in v["features"]:
            if key not in FIXED_FEATURE_KEYS and key not in seen_extra:
                seen_extra.add(key)
                extra_keys.append(key)

    groups = {}
    for v in variants:
        series_key = v["features"].get("Серия", v.get("source_url", v["url"]))
        groups.setdefault(series_key, []).append(v)

    base_ts = int(datetime.now().timestamp() * 1000)

    # Шаблон Яндекс Кит — столбцы строго по порядку из официального шаблона
    fieldnames = [
        "KIT ID*", "Название товара*", "Артикул", "Описание товара", "Штрихкод",
        "Статус", "Цена до скидки, руб.", "Цена со скидкой, руб.", "Цена по акции, руб.",
        "Минимальная цена, руб.", "Приоритет в каталоге", "Ставка НДС", "Маркируемый товар",
        "Объединять по (Group ID)", "Группирующие характеристики", "Разделять на карточки по",
        "Бренд",
        "Категория 1-го уровня*", "Категория 2-го уровня", "Категория 3-го уровня",
        "Склад: Склад №1",
        "Внешний ID: 1C", "Внешний ID: МойСклад", "Внешний ID: Wildberries",
        "Внешний ID: Ozon", "Внешний ID: YML", "Внешний ID: Яндекс.Маркет",
        "Изображения и видео", "Ссылки на файлы", "Бейджи",
        "Количество упаковок", "Высота упаковки, см", "Ширина упаковки, см",
        "Длина упаковки, см", "Вес с упаковкой, г",
        # Характеристики
        "Цвет", "Объём встроенной памяти", "Тип связи",
        "Серия", "Процессор", "Диагональ",
        "Основная камера", "Фронтальная камера",
        "ОС", "Защита от воды", "Разъём", "Год модели",
    ]
    fieldnames.extend(extra_keys)

    # Map XLSX column names → feature dict keys
    FEATURE_KEY_MAP = {
        "Цвет": "Цвет",
        "Объём встроенной памяти": "Встроенная память",
        "Тип связи": "Связь",
        "Процессор": "Процессор",
    }

    def _make_grouping_key(f, grouping_cols):
        return "|".join(f.get(FEATURE_KEY_MAP.get(c, c), "") for c in grouping_cols)

    def _pick_grouping(group_variants):
        all_have = lambda key: all(v["features"].get(key, "") for v in group_variants)
        _is_macbook = any(
            "macbook" in (v.get("name", "") + " " + v.get("url", "")).lower()
            or "mac" in v["features"].get("Тип", "").lower()
            for v in group_variants
        )
        if all_have("Связь"):
            return ["Цвет", "Объём встроенной памяти", "Тип связи"]
        if _is_macbook:
            base = ["Цвет", "Объём встроенной памяти"]
            if all_have("Оперативная память"):
                base.append("Оперативная память")
            return base
        if all_have("Процессор"):
            processors = {v["features"].get("Процессор", "") for v in group_variants}
            if len(processors) > 1:
                return ["Цвет", "Объём встроенной памяти", "Процессор"]
        return ["Цвет", "Объём встроенной памяти"]

    def _category(v):
        name = (v.get("name", "") + " " + v["features"].get("Серия", "")).lower()
        url = v.get("url", "").lower()
        if "macbook" in name or "macbook" in url or "mac" in v["features"].get("Тип", "").lower():
            return ("Электроника", "Ноутбуки", "Apple MacBook")
        if "ipad" in name or "ipad" in url:
            return ("Электроника", "Планшеты", "Apple iPad")
        if "airpods" in name or "airpods" in url:
            return ("Электроника", "Наушники", "Apple AirPods")
        if "watch" in url or "apple watch" in name:
            return ("Электроника", "Смарт-часы", "Apple Watch")
        return ("Электроника", "Смартфоны", "Apple iPhone")

    wb = openpyxl.Workbook()
    ws = wb.active

    # header row
    for ci, h in enumerate(fieldnames, 1):
        ws.cell(row=1, column=ci, value=h)

    ri = 2
    for gi, (src_url, group_variants) in enumerate(groups.items()):
        group_id = str(base_ts + gi)
        series = group_variants[0]["features"].get("Серия", "")

        grouping_cols = _pick_grouping(group_variants)
        grouping_chars_str = "; ".join(grouping_cols)

        # Deduplicate: skip only truly identical variants (same features)
        seen_hashes = {}
        deduped = []
        skipped = 0
        for v in group_variants:
            f = v["features"]
            dedup_key = "|".join(f"{k}={v}" for k, v in sorted(f.items()) if v)
            if dedup_key in seen_hashes:
                skipped += 1
                continue
            seen_hashes[dedup_key] = True
            deduped.append(v)
        if skipped:
            print(f"  [!] Пропущено дубликатов: {skipped} (группа {series})")

        for v in deduped:
            f = v["features"]
            old_price = int(v["old_price"]) if v["old_price"] > 0 else int(v["price"])
            weight_raw = f.get("Вес", "").replace(" г", "").strip()
            row = [
                "",                             # KIT ID* — пусто для новых
                _build_product_name(v, f, series),  # Название товара*
                v["sku"],                       # Артикул
                "",                             # Описание товара
                "",                             # Штрихкод
                "Опубликован",                  # Статус
                old_price,                      # Цена до скидки, руб.
                int(v["price"]),                # Цена со скидкой, руб.
                "",                             # Цена по акции, руб.
                "",                             # Минимальная цена, руб.
                "",                             # Приоритет в каталоге
                22,                             # Ставка НДС
                0,                              # Маркируемый товар
                group_id,                       # Объединять по (Group ID)
                grouping_chars_str,             # Группирующие характеристики
                "",                             # Разделять на карточки по
                f.get("Бренд", "Apple"),  # Бренд
                _category(v)[0],                # Категория 1-го уровня*
                _category(v)[1],                # Категория 2-го уровня
                _category(v)[2],                # Категория 3-го уровня
                v["stock"],                     # Склад: Склад №1
                "", "", "",                     # Внешние ID (1C, МойСклад, WB)
                "", "", "",                     # Внешние ID (Ozon, YML, Яндекс.Маркет)
                " ".join(v.get("images", [])[:10]),  # Изображения и видео
                "",                             # Ссылки на файлы
                "",                             # Бейджи
                "",                             # Количество упаковок
                "", "", "",                     # Габариты (высота, ширина, длина)
                weight_raw,                     # Вес с упаковкой, г
                # Характеристики
                f.get("Цвет", ""),
                f.get("Встроенная память", ""),
                f.get("Связь", ""),
                f.get("Серия", ""),
                f.get("Процессор", ""),
                f.get("Диагональ", ""),
                f.get("Разрешение камеры", ""),
                f.get("Разрешение фронтальной камеры, Мп", ""),
                f.get("Операционная система", ""),
                f.get("Защита от воды", ""),
                f.get("Разъём", ""),
                f.get("Год (модель представлена)", ""),
            ]
            # Extra characteristic columns
            row.extend(f.get(key, "") for key in extra_keys)
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
            ri += 1

    wb.save(output_path)
    return ri - 2


# ====================================================================
# YANDEX MARKET YML GENERATION
# ====================================================================

def _feature_val(features, *keys):
    for k in keys:
        v = features.get(k, "")
        if v:
            return v
    return ""


def generate_yandex_market_yml(variants, output_path):

    if not variants:
        return 0

    groups = {}
    for v in variants:
        series_key = v["features"].get("Серия", v.get("source_url", v["url"]))
        groups.setdefault(series_key, []).append(v)

    base_ts = int(datetime.now().timestamp() * 1000)
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S+03:00")

    yml = Element("yml_catalog", {"date": now})
    shop = SubElement(yml, "shop")

    SubElement(shop, "name").text = "Apple"
    SubElement(shop, "company").text = "Apple"
    SubElement(shop, "url").text = "https://www.apple.com"

    currencies = SubElement(shop, "currencies")
    SubElement(currencies, "currency", {"id": "RUB", "rate": "1"})

    categories = SubElement(shop, "categories")
    SubElement(categories, "category", {"id": "1"}).text = "Смартфоны"
    SubElement(categories, "category", {"id": "2", "parentId": "1"}).text = "Apple iPhone"
    SubElement(categories, "category", {"id": "3"}).text = "Ноутбуки"
    SubElement(categories, "category", {"id": "4", "parentId": "3"}).text = "Apple MacBook"
    SubElement(categories, "category", {"id": "5"}).text = "Планшеты"
    SubElement(categories, "category", {"id": "6", "parentId": "5"}).text = "Apple iPad"
    first_series = list(groups.values())[0][0]["features"].get("Серия", "iPhone")
    SubElement(categories, "category", {"id": "7", "parentId": "2"}).text = f"Apple {first_series}"

    offers = SubElement(shop, "offers")

    def _yml_category(v):
        name = (v.get("name", "") + " " + v["features"].get("Серия", "")).lower()
        url = v.get("url", "").lower()
        if "macbook" in name or "macbook" in url or "mac" in v["features"].get("Тип", "").lower():
            return "4"
        if "ipad" in name or "ipad" in url:
            return "6"
        return "2"

    for gi, (src_url, group_variants) in enumerate(groups.items()):
        group_id = str(base_ts + gi)

        for v in group_variants:
            f = v["features"]
            offer = SubElement(offers, "offer", {
                "id": v["sku"],
                "available": "true",
                "group_id": group_id,
            })

            SubElement(offer, "url").text = v["url"]
            SubElement(offer, "price").text = str(int(v["price"]))
            old_price = int(v["old_price"]) if v["old_price"] > 0 else int(v["price"])
            SubElement(offer, "oldprice").text = str(old_price)
            SubElement(offer, "currencyId").text = "RUB"
            SubElement(offer, "categoryId").text = _yml_category(v)

            for img_url in v["images"][:10]:
                SubElement(offer, "picture").text = img_url

            name = _feature_val(f, "Серия")
            SubElement(offer, "name").text = name if name else v["name"]

            SubElement(offer, "vendor").text = f.get("Бренд", "Apple")
            SubElement(offer, "vendorCode").text = v["sku"]

            SubElement(offer, "description").text = ""

            SubElement(offer, "barcode")
            weight_val = f.get("Вес", "").replace(" г", "").strip()
            SubElement(offer, "weight").text = weight_val if weight_val else "0"

            for fkey, fval in f.items():
                if fval:
                    SubElement(offer, "param", {"name": fkey}).text = str(fval)

    raw = tostring(yml, encoding="unicode")
    pretty = parseString(raw).toprettyxml(indent="  ", encoding="UTF-8")
    with open(output_path, "wb") as f:
        f.write(pretty)

    return len(variants)


# ====================================================================
# MAIN
# ====================================================================

def main():
    max_workers = 10
    args = [a for a in sys.argv[1:] if not a.startswith("--workers=")]
    for a in sys.argv[1:]:
        if a.startswith("--workers="):
            max_workers = int(a.split("=")[1])

    if len(args) >= 2 and args[0] == "--file":
        variants = parse_urls_from_file(args[1], max_workers=max_workers)
    elif len(args) > 0 and not args[0].startswith("--"):
        variants = parse_product(args[0])
    else:
        variants = parse_urls_from_file("urls.txt", max_workers=max_workers)

    print(f"\n{'=' * 60}")
    print(f"Найдено вариаций: {len(variants)}")
    print(f"{'=' * 60}")

    for v in variants:
        f = v["features"]
        color = f.get("Цвет", "?")
        memory = f.get("Встроенная память", "?")
        conn = f.get("Связь", "?")
        print(f"  {v['sku']:12s} | {color:12s} | {memory:8s} | {conn:14s} | {int(v['price']):>8,} RUB | stock: {v['stock']}".replace(",", " "))

    # Группируем по категориям — каждая в отдельный файл
    by_category = {}
    for v in variants:
        cat = detect_category(v.get("url", ""))
        by_category.setdefault(cat, []).append(v)

    for cat_name, cat_variants in by_category.items():
        print(f"\n--- Категория: {cat_name} ({len(cat_variants)} вариаций) ---")

        xlsx_path = f"yandex-kit-import-{cat_name}.xlsx"
        n = generate_yandex_kit_xlsx(cat_variants, xlsx_path)
        print(f"  XLSX: {xlsx_path} ({n} строк)")

        yml_path = f"yandex-market-import-{cat_name}.yml"
        n2 = generate_yandex_market_yml(cat_variants, yml_path)
        print(f"  YML:  {yml_path} ({n2} офферов)")

        json_path = f"parsed-data-{cat_name}.json"
        serializable = []
        for v in cat_variants:
            serializable.append({
                "sku": v["sku"],
                "name": v["name"],
                "price": v["price"],
                "old_price": v["old_price"],
                "stock": v["stock"],
                "url": v["url"],
                "source_url": v.get("source_url", v["url"]),
                "brand": v["features"].get("Бренд", "Apple"),
                "description": v.get("description", ""),
                "features": v["features"],
                "images": v.get("images", []),
            })
        with open(json_path, "w", encoding="utf-8") as fh:
            json.dump({"variants": serializable}, fh, ensure_ascii=False, indent=2)
        print(f"  JSON: {json_path}")

    print("\nГотово!")
    print("  Загрузка в Яндекс Кит:   Товары -> Добавить -> из Excel-файла -> yandex-kit-import-*.xlsx")
    print("  Загрузка в Яндекс Маркет: yandex-market-import-*.yml")


if __name__ == "__main__":
    main()
